from io import BytesIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

_FILL_HEADER = PatternFill("solid", fgColor="1e293b")   # slate-800
_FILL_ALT    = PatternFill("solid", fgColor="f1f5f9")   # slate-100
_FONT_HEADER = Font(bold=True, color="FFFFFF", size=10)
_FONT_BODY   = Font(size=10)
_BORDER_THIN = Border(
    bottom=Side(style="thin", color="CBD5E1"),
)

MESES = [
    "", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
    "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre",
]

ESTADO_LABEL = {
    "operativo":        "Operativo",
    "en_mantenimiento": "En mantenimiento",
    "fuera_servicio":   "Fuera de servicio",
    "en_traslado":      "En traslado",
    "baja":             "Baja",
}


def _auto_width(ws, extra=4):
    for col_cells in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col_cells), default=0)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + extra, 50)


def _write_header_row(ws, row_num, cols):
    for ci, text in enumerate(cols, start=1):
        cell = ws.cell(row=row_num, column=ci, value=text)
        cell.font = _FONT_HEADER
        cell.fill = _FILL_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _write_data_row(ws, row_num, values, alt=False):
    fill = _FILL_ALT if alt else None
    for ci, val in enumerate(values, start=1):
        cell = ws.cell(row=row_num, column=ci, value=val)
        cell.font = _FONT_BODY
        cell.border = _BORDER_THIN
        if fill:
            cell.fill = fill


# ─────────────────────────────────────────────────────────────
# Reporte Diario
# ─────────────────────────────────────────────────────────────

def generar_excel_reporte_diario(datos: dict, criticos: list) -> bytes:
    wb = Workbook()

    # ── Hoja 1: Métricas ──
    ws1 = wb.active
    ws1.title = "Resumen del día"

    hoy = datos.get("fecha", datetime.now().strftime("%Y-%m-%d"))
    ws1["A1"] = f"Reporte Diario — {hoy}"
    ws1["A1"].font = Font(bold=True, size=13)
    ws1["A2"] = "IMSS — HGR No. 1 Tijuana"
    ws1["A2"].font = Font(italic=True, size=10, color="64748B")
    ws1.append([])

    estado_map = {e["estado"]: e["total"] for e in (datos.get("equipos_por_estado") or [])}
    metricas = [
        ("Órdenes abiertas hoy",     datos.get("ordenes_hoy", 0)),
        ("Órdenes pendientes",       datos.get("ordenes_abiertas", 0)),
        ("Equipos operativos",       estado_map.get("operativo", 0)),
        ("En mantenimiento",         estado_map.get("en_mantenimiento", 0)),
        ("Fuera de servicio",        estado_map.get("fuera_servicio", 0)),
        ("En traslado",              estado_map.get("en_traslado", 0)),
    ]
    _write_header_row(ws1, 4, ["Indicador", "Total"])
    for i, (label, val) in enumerate(metricas, start=5):
        _write_data_row(ws1, i, [label, val], alt=(i % 2 == 0))

    # ── Sección: preventivos ──
    preventivos = datos.get("preventivos_proxima_semana") or []
    if preventivos:
        start = 5 + len(metricas) + 2
        ws1.cell(row=start - 1, column=1, value="Preventivos próximos 7 días").font = Font(bold=True, size=11)
        _write_header_row(ws1, start, ["Equipo", "Serie", "Tipo preventivo", "Fecha"])
        for i, pp in enumerate(preventivos, start=start + 1):
            _write_data_row(ws1, i, [pp["nombre"], pp["serie"], pp["tipo_preventivo"], str(pp["proxima_ejecucion"])], alt=(i % 2 == 0))

    _auto_width(ws1)

    # ── Hoja 2: Equipos críticos ──
    if criticos:
        ws2 = wb.create_sheet("Equipos críticos")
        ws2["A1"] = "Equipos críticos / fuera de servicio"
        ws2["A1"].font = Font(bold=True, size=13)
        ws2.append([])
        _write_header_row(ws2, 3, ["Equipo", "Marca", "Modelo", "Serie", "Estado", "Área", "Tickets abiertos"])
        for i, eq in enumerate(criticos, start=4):
            estado = ESTADO_LABEL.get(eq.get("estado", ""), eq.get("estado", ""))
            _write_data_row(ws2, i, [
                eq.get("nombre"), eq.get("marca"), eq.get("modelo"),
                eq.get("serie"), estado, eq.get("area"), eq.get("tickets_abiertos", 0),
            ], alt=(i % 2 == 0))
        _auto_width(ws2)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────
# Historial de Órdenes
# ─────────────────────────────────────────────────────────────

def generar_excel_historial(mes: int, anio: int, ordenes: list) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = f"{MESES[mes]} {anio}"

    ws["A1"] = f"Historial de Órdenes — {MESES[mes]} {anio}"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"] = "IMSS — HGR No. 1 Tijuana"
    ws["A2"].font = Font(italic=True, size=10, color="64748B")
    ws.append([])

    cols = ["No. Orden", "Equipo", "Tipo mantenimiento", "Estado", "Fecha", "Área", "Técnico"]
    _write_header_row(ws, 4, cols)

    for i, o in enumerate(ordenes, start=5):
        fecha = str(o.get("fecha", ""))[:10]
        _write_data_row(ws, i, [
            o.get("numero_orden", o.get("id", "")),
            o.get("equipo_nombre_rel", ""),
            o.get("tipo_mantenimiento", ""),
            o.get("estado", ""),
            fecha,
            o.get("area", ""),
            o.get("tecnico_nombre", ""),
        ], alt=(i % 2 == 0))

    _auto_width(ws)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
