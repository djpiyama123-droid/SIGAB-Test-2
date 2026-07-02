#!/usr/bin/env python3
"""
descargar_751_y_armar_excel.py
================================================================
Descarga los 751 equipos de SIGAB prod (read-only) y los pone en una
plantilla Excel pre-llenada para que Gustavo solo concilie en el hospital,
no tenga que transcribir desde cero.

USO:
    python3 descargar_751_y_armar_excel.py --token <JWT>

El token se obtiene de /api/auth/login en prod:
    curl -X POST https://sigab.129-121-100-147.sslip.io/api/auth/login \
      -H 'Content-Type: application/json' \
      -d '{"matricula":"ADMIN001","password":"sigah_admin_2026"}'

SALIDA:
    piloto_hgr1/levantamiento_2026-06.xlsx — Excel con 5 hojas
    piloto_hgr1/equipos_backup_2026-06-23.json — JSON crudo por si acaso
    piloto_hgr1/estado_descarga.txt — log de la corrida
"""
import argparse
import json
import sys
from datetime import datetime
from pathlib import Path
import requests
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, NamedStyle
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

API_BASE = "https://sigab.129-121-100-147.sslip.io"
OUT_DIR = Path("/mnt/c/Users/djpiy/Desktop/Bioingeneria/SIGAB/piloto_hgr1")

# -------------------------------------------------------------------
# 1. DESCARGA DE PROD
# -------------------------------------------------------------------
def descargar_equipos(token: str) -> list[dict]:
    """Descarga TODOS los equipos paginando. Read-only contra prod."""
    headers = {"Authorization": f"Bearer {token}"}
    todos = []
    offset = 0
    limit = 200  # máximo razonable
    pagina = 0

    while True:
        pagina += 1
        url = f"{API_BASE}/api/equipos/?limit={limit}&offset={offset}"
        r = requests.get(url, headers=headers, timeout=30)
        if r.status_code == 401:
            sys.exit("❌ Token inválido o expirado. Vuelve a hacer login.")
        r.raise_for_status()
        data = r.json()
        equipos = data.get("equipos", [])
        total = data.get("total", 0)
        todos.extend(equipos)
        print(f"  página {pagina}: {len(equipos)} equipos (acumulado: {len(todos)}/{total})")
        if len(todos) >= total or not equipos:
            break
        offset += limit

    return todos


# -------------------------------------------------------------------
# 2. ARMADO DEL EXCEL
# -------------------------------------------------------------------
COLUMNAS = [
    # (header, ancho, campo_json, tipo, requerido)
    ("#",                      5,   None,                "int", False),
    ("id_equipo [auto]",       12,  "id",                "auto", False),
    ("codigo_inventario",      18,  "inventario",        "str", True),
    ("numero_serie",           22,  "serie",             "str", True),
    ("nombre",                 32,  "nombre",            "str", True),
    ("marca",                  18,  "marca",             "str", True),
    ("modelo",                 22,  "modelo",            "str", True),
    ("piso",                   8,   "piso",              "int", True),
    ("área",                   16,  "area",              "str", True),
    ("ubicacion_especifica",   32,  "ubicacion",         "str", False),
    ("estado",                 16,  "estado",            "enum", True),
    ("criticidad",             12,  "criticidad",        "enum", False),
    ("tipo_equipo",            18,  "tipo_equipo",       "enum", False),
    ("clase_cofepris",         14,  "clase_cofepris",    "enum", False),
    ("fecha_instalacion",      16,  "fecha_instalacion", "date", False),
    ("fecha_ultimo_preventivo",18,  "fecha_ultimo_mantenimiento", "date", False),
    ("numero_contrato",        18,  "numero_contrato",   "str", False),
    ("proveedor_servicio",     22,  "proveedor_servicio","str", False),
    ("observaciones_campo",    36,  None,                "str", False),
    ("qr_token [auto]",        22,  "qr_token",          "auto", False),
    ("conciliado_sigab",       16,  None,                "bool", True),
    ("estaba_en_piso",         16,  None,                "bool", False),
    ("foto_url",               30,  "imagen_url",        "str", False),
]

ESTADOS_VALIDOS = ["operativo", "mantenimiento", "fuera_servicio", "baja", "en_traslado"]
CRITICIDAD_VALIDOS = ["alta", "media", "baja"]
TIPOS_VALIDOS = [
    "ventilador", "monitor", "bomba", "desfibrilador", "ultrasonido",
    "rayos_x", "electrocardiografo", "anestesia", "incubadora",
    "otro", "monitor_fetal", "oximetro", " aspirador"
]
COFEPRIS_VALIDOS = ["I", "II", "III"]


def _apply_border(ws, cell_range):
    """Bordes finos a un rango."""
    side = Side(border_style="thin", color="d0d7de")
    border = Border(left=side, right=side, top=side, bottom=side)
    for row in ws[cell_range]:
        for cell in row:
            cell.border = border


def _color_por_estado(valor: str) -> str:
    return {
        "operativo":      "c6f6d5",  # verde claro
        "mantenimiento":  "fef3c7",  # amarillo claro
        "fuera_servicio": "fed7d7",  # rojo claro
        "baja":           "e2e8f0",  # gris claro
        "en_traslado":    "dbeafe",  # azul claro
    }.get(valor, "ffffff")


def _crear_hoja_levantamiento(wb, equipos: list[dict]) -> None:
    """Hoja principal: 751 filas pre-llenadas + columnas extras para captura.
    Se inserta en la posición 1 (detrás de Instrucciones)."""
    ws = wb.create_sheet("Levantamiento")  # al final, se reordena al final

    # --- Header ---
    header_fill = PatternFill("solid", fgColor="006CB7")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for col_idx, (header, ancho, *_ ) in enumerate(COLUMNAS, start=1):
        c = ws.cell(row=1, column=col_idx, value=header)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = ancho
    ws.row_dimensions[1].height = 36

    # --- Filas de datos ---
    for i, eq in enumerate(equipos, start=1):
        row = i + 1
        # # correlativo
        ws.cell(row=row, column=1, value=i)
        # id_equipo (read-only gris)
        c = ws.cell(row=row, column=2, value=eq.get("id"))
        c.font = Font(color="6a737d", italic=True)
        # codigo_inventario
        ws.cell(row=row, column=3, value=eq.get("inventario") or "")
        # numero_serie
        ws.cell(row=row, column=4, value=eq.get("serie") or "")
        # nombre
        ws.cell(row=row, column=5, value=eq.get("nombre") or "")
        # marca
        ws.cell(row=row, column=6, value=eq.get("marca") or "")
        # modelo
        ws.cell(row=row, column=7, value=eq.get("modelo") or "")
        # piso
        ws.cell(row=row, column=8, value=eq.get("piso") or "")
        # area
        ws.cell(row=row, column=9, value=eq.get("area") or "")
        # ubicacion
        ws.cell(row=row, column=10, value=eq.get("ubicacion") or "")
        # estado con color
        estado = eq.get("estado") or "operativo"
        c = ws.cell(row=row, column=11, value=estado)
        c.fill = PatternFill("solid", fgColor=_color_por_estado(estado))
        # criticidad
        ws.cell(row=row, column=12, value=eq.get("criticidad") or "media")
        # tipo_equipo
        ws.cell(row=row, column=13, value=eq.get("tipo_equipo") or "otro")
        # clase_cofepris
        ws.cell(row=row, column=14, value=eq.get("clase_cofepris") or "II")
        # fechas
        fi = eq.get("fecha_instalacion")
        ws.cell(row=row, column=15, value=str(fi) if fi else "")
        fup = eq.get("fecha_ultimo_mantenimiento")
        ws.cell(row=row, column=16, value=str(fup) if fup else "")
        # contrato/proveedor
        ws.cell(row=row, column=17, value=eq.get("numero_contrato") or "")
        ws.cell(row=row, column=18, value=eq.get("proveedor_servicio") or "")
        # observaciones_campo (vacío — para que Gustavo anote)
        # qr_token (read-only)
        c = ws.cell(row=row, column=20, value=eq.get("qr_token") or "")
        c.font = Font(name="Consolas", color="006CB7", size=9)
        # conciliado_sigab default SÍ (porque ya está en prod)
        c = ws.cell(row=row, column=21, value="SÍ")
        c.font = Font(bold=True, color="22863a")
        c.alignment = Alignment(horizontal="center")
        # estaba_en_piso default VERDE (para que Gustavo marque ✗ si no aparece)
        c = ws.cell(row=row, column=22, value="✓")
        c.font = Font(bold=True, color="22863a", size=14)
        c.alignment = Alignment(horizontal="center")
        # foto_url
        ws.cell(row=row, column=23, value=eq.get("imagen_url") or "")

    last_row = len(equipos) + 1

    # --- Validaciones (dropdowns) ---
    # Estado
    dv_estado = DataValidation(
        type="list",
        formula1='"' + ",".join(ESTADOS_VALIDOS) + '"',
        allow_blank=False,
    )
    dv_estado.error = "Estado inválido"
    dv_estado.add(f"K2:K{last_row}")
    ws.add_data_validation(dv_estado)
    # Criticidad
    dv_crit = DataValidation(
        type="list",
        formula1='"' + ",".join(CRITICIDAD_VALIDOS) + '"',
        allow_blank=True,
    )
    dv_crit.add(f"L2:L{last_row}")
    ws.add_data_validation(dv_crit)
    # Tipo
    dv_tipo = DataValidation(
        type="list",
        formula1='"' + ",".join(TIPOS_VALIDOS) + '"',
        allow_blank=True,
    )
    dv_tipo.add(f"M2:M{last_row}")
    ws.add_data_validation(dv_tipo)
    # COFEPRIS
    dv_cof = DataValidation(
        type="list",
        formula1='"' + ",".join(COFEPRIS_VALIDOS) + '"',
        allow_blank=True,
    )
    dv_cof.add(f"N2:N{last_row}")
    ws.add_data_validation(dv_cof)
    # Conciliado
    dv_conc = DataValidation(
        type="list", formula1='"SÍ,NO,PENDIENTE"', allow_blank=False,
    )
    dv_conc.add(f"U2:U{last_row}")
    ws.add_data_validation(dv_conc)
    # Estaba en piso
    dv_piso = DataValidation(
        type="list", formula1='"✓,✗,?", ', allow_blank=False,
    )
    dv_piso.add(f"V2:V{last_row}")
    ws.add_data_validation(dv_piso)

    # --- Freeze header + filtros ---
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNAS))}{last_row}"

    # Bordes
    _apply_border(ws, f"A1:{get_column_letter(len(COLUMNAS))}{last_row}")


def _crear_hoja_instrucciones(wb) -> None:
    """Hoja 0: cómo usar este Excel (la primera que ve Gustavo al abrir)."""
    ws = wb.create_sheet("Instrucciones")  # al inicio
    instrucciones = [
        ("LEVANTAMIENTO DE EQUIPOS BIOMÉDICOS — HGR No.1 IMSS Tijuana", "titulo"),
        ("Semana 23-27 junio 2026 · Operador: Gustavo López Carballo", "subtitulo"),
        ("", ""),
        ("COLUMNAS:", "h2"),
        ("  • # — correlativo (no tocar, lo pone la hoja)", "li"),
        ("  • id_equipo — lo asigna SIGAB, no tocar", "li"),
        ("  • codigo_inventario — código interno del hospital (ej. HGR-EC-001)", "li"),
        ("  • numero_serie — serie del fabricante (o código interno si no hay)", "li"),
        ("  • nombre / marca / modelo — datos del equipo", "li"),
        ("  • piso / área / ubicacion_especifica — dónde está FÍSICAMENTE", "li"),
        ("  • estado — operativo / mantenimiento / fuera_servicio / baja / en_traslado", "li"),
        ("  • criticidad — alta / media / baja", "li"),
        ("  • tipo_equipo — dropdown con valores predefinidos", "li"),
        ("  • clase_cofepris — I / II / III", "li"),
        ("  • fecha_instalacion / fecha_ultimo_preventivo — formato YYYY-MM-DD", "li"),
        ("  • numero_contrato / proveedor_servicio — datos administrativos", "li"),
        ("  • observaciones_campo — TUS NOTAS del recorrido (NO lo que ya estaba en SIGAB)", "li"),
        ("  • qr_token — NO TOCAR, ya está generado por SIGAB prod", "li"),
        ("  • conciliado_sigab — SÍ (ya está) / NO (no aparece en físico) / PENDIENTE", "li"),
        ("  • estaba_en_piso — ✓ SÍ / ✗ NO / ? DUDOSO (lo llenas TÚ en el hospital)", "li"),
        ("  • foto_url — link a la foto del equipo en SIGAB (si existe)", "li"),
        ("", ""),
        ("REGLAS DE ORO:", "h2"),
        ("  1. Empieza por las áreas críticas: UCI, Urgencias, Quirófano, Imagenología.", "li"),
        ("  2. Lleva la tablet o laptop. Si no, libreta y pasa al Excel al final del día.", "li"),
        ("  3. serie es CLAVE ÚNICA. Si no se lee, asigna HGR-EC-NXX.", "li"),
        ("  4. Si un equipo NO está en SIGAB prod, márcalo conciliado=NO y explica en observaciones.", "li"),
        ("  5. Si un equipo SÍ está pero no lo encuentras, márcalo estaba_en_piso=✗ y notifícale a Gustavo.", "li"),
        ("  6. NO borres filas. Si un equipo ya no existe, márcalo estado=baja y conciliado=NO.", "li"),
        ("", ""),
        ("DESPUÉS DEL RECORRIDO:", "h2"),
        ("  → Renombra el archivo a levantamiento_2026-06_FINAL.xlsx", "li"),
        ("  → Avísale a Hermes: el script de importación批量 sube los nuevos a SIGAB.", "li"),
        ("", ""),
        ("LEYENDA DE COLORES:", "h2"),
        ("  • VERDE = conciliado / OK", "li"),
        ("  • AMARILLO = en mantenimiento", "li"),
        ("  • ROJO = fuera de servicio / falta conciliar", "li"),
        ("  • GRIS = dado de baja", "li"),
        ("  • AZUL CLARO = en traslado", "li"),
    ]
    for fila, (texto, estilo) in enumerate(instrucciones, start=1):
        c = ws.cell(row=fila, column=1, value=texto)
        if estilo == "titulo":
            c.font = Font(bold=True, size=16, color="006CB7")
        elif estilo == "subtitulo":
            c.font = Font(italic=True, size=11, color="6a737d")
        elif estilo == "h2":
            c.font = Font(bold=True, size=12, color="1a1a1a")
        # column A ancho
        ws.column_dimensions["A"].width = 80


def _crear_hoja_stats(wb, equipos: list[dict]) -> None:
    """Hoja de stats: por piso, por estado, por área, por tipo."""
    ws = wb.create_sheet("Stats")
    ws["A1"] = "ESTADÍSTICAS DEL INVENTARIO"
    ws["A1"].font = Font(bold=True, size=14, color="006CB7")

    # Por piso
    ws["A3"] = "Por piso"
    ws["A3"].font = Font(bold=True)
    pisos = {}
    for eq in equipos:
        p = str(eq.get("piso") or "—")
        pisos[p] = pisos.get(p, 0) + 1
    for i, (p, n) in enumerate(sorted(pisos.items()), start=4):
        ws.cell(row=i, column=1, value=p)
        ws.cell(row=i, column=2, value=n)

    # Por estado
    col_offset = 5
    ws.cell(row=3, column=col_offset, value="Por estado").font = Font(bold=True)
    estados = {}
    for eq in equipos:
        e = eq.get("estado") or "—"
        estados[e] = estados.get(e, 0) + 1
    for i, (e, n) in enumerate(sorted(estados.items()), start=4):
        ws.cell(row=i, column=col_offset, value=e).fill = PatternFill(
            "solid", fgColor=_color_por_estado(e)
        )
        ws.cell(row=i, column=col_offset+1, value=n)

    # Por área
    col_offset = 9
    ws.cell(row=3, column=col_offset, value="Por área").font = Font(bold=True)
    areas = {}
    for eq in equipos:
        a = eq.get("area") or "—"
        areas[a] = areas.get(a, 0) + 1
    for i, (a, n) in enumerate(sorted(areas.items()), start=4):
        ws.cell(row=i, column=col_offset, value=a)
        ws.cell(row=i, column=col_offset+1, value=n)

    # Por tipo
    col_offset = 13
    ws.cell(row=3, column=col_offset, value="Por tipo_equipo").font = Font(bold=True)
    tipos = {}
    for eq in equipos:
        t = eq.get("tipo_equipo") or "—"
        tipos[t] = tipos.get(t, 0) + 1
    for i, (t, n) in enumerate(sorted(tipos.items()), start=4):
        ws.cell(row=i, column=col_offset, value=t)
        ws.cell(row=i, column=col_offset+1, value=n)

    for col in [1, 5, 9, 13]:
        ws.column_dimensions[get_column_letter(col)].width = 22
    for col in [2, 6, 10, 14]:
        ws.column_dimensions[get_column_letter(col)].width = 8


def _crear_hoja_pendientes(wb) -> None:
    """Hoja 'Pendientes' para que Gustavo anote equipos NUEVOS no en prod.
    Es la lista que se subirá con el importador Excel."""
    ws = wb.create_sheet("Pendientes_Nuevos")  # al final, se reordena
    headers = [
        ("#", 5), ("codigo_inventario", 18), ("numero_serie", 22),
        ("nombre", 32), ("marca", 18), ("modelo", 22), ("piso", 8),
        ("area", 16), ("ubicacion_especifica", 32), ("estado", 16),
        ("criticidad", 12), ("tipo_equipo", 18), ("clase_cofepris", 14),
        ("fecha_instalacion", 16), ("numero_contrato", 18),
        ("proveedor_servicio", 22), ("observaciones", 36),
    ]
    for col_idx, (h, w) in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.fill = PatternFill("solid", fgColor="006CB7")
        c.font = Font(bold=True, color="FFFFFF")
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    # Fila de ejemplo
    ws.cell(row=2, column=1, value=1)
    ws.cell(row=2, column=2, value="HGR-EC-N01")
    ws.cell(row=2, column=3, value="GE-VS-2024-A891")
    ws.cell(row=2, column=4, value="Ventilador de transporte")
    ws.cell(row=2, column=5, value="GE Healthcare")
    ws.cell(row=2, column=6, value="Carescape R860")
    ws.cell(row=2, column=7, value=1)
    ws.cell(row=2, column=8, value="Urgencias")
    ws.cell(row=2, column=9, value="P1, Urg, shock-trauma")
    ws.cell(row=2, column=10, value="operativo")
    ws.cell(row=2, column=11, value="alta")
    ws.cell(row=2, column=12, value="ventilador")
    ws.cell(row=2, column=13, value="III")
    ws.cell(row=2, column=14, value="2024-11-08")
    ws.cell(row=2, column=15, value="CONT-2024-156")
    ws.cell(row=2, column=16, value="GE Healthcare")
    ws.cell(row=2, column=17, value="Equipo nuevo no registrado en SIGAB prod")

    _apply_border(ws, f"A1:{get_column_letter(len(headers))}2")
    ws.freeze_panes = "A2"


# -------------------------------------------------------------------
# 3. MAIN
# -------------------------------------------------------------------
def main():
    p = argparse.ArgumentParser()
    p.add_argument("--token", required=True, help="JWT token de SIGAB prod")
    p.add_argument("--out", default=str(OUT_DIR), help="directorio de salida")
    args = p.parse_args()

    out = Path(args.out)
    out.mkdir(parents=True, exist_ok=True)

    print(f"📥 Descargando equipos de {API_BASE} ...")
    equipos = descargar_equipos(args.token)
    print(f"✅ {len(equipos)} equipos descargados")

    # Guardar JSON crudo
    json_path = out / f"equipos_backup_{datetime.now().strftime('%Y-%m-%d')}.json"
    json_path.write_text(json.dumps(equipos, indent=2, default=str))
    print(f"💾 Backup JSON: {json_path}")

    # Armar Excel
    print("📊 Armando Excel ...")
    wb = Workbook()
    wb.remove(wb.active)  # quitar hoja por default
    _crear_hoja_instrucciones(wb)
    _crear_hoja_levantamiento(wb, equipos)
    _crear_hoja_stats(wb, equipos)
    _crear_hoja_pendientes(wb)

    # Reordenar: Instrucciones, Levantamiento, Stats, Pendientes_Nuevos
    orden = ["Instrucciones", "Levantamiento", "Stats", "Pendientes_Nuevos"]
    wb._sheets = [wb[name] for name in orden if name in wb.sheetnames]

    # Hacer que "Levantamiento" sea la hoja activa al abrir
    wb.active = wb.sheetnames.index("Levantamiento")

    xlsx_path = out / "levantamiento_2026-06.xlsx"
    wb.save(xlsx_path)
    print(f"✅ Excel: {xlsx_path}")

    # Log
    log_path = out / "estado_descarga.txt"
    log_path.write_text(
        f"Descarga: {datetime.now().isoformat()}\n"
        f"Equipos: {len(equipos)}\n"
        f"API: {API_BASE}\n"
        f"Excel: {xlsx_path}\n"
    )
    print(f"📝 Log: {log_path}")
    print()
    print("=" * 60)
    print("PRÓXIMOS PASOS:")
    print("  1. Abre el Excel en tu laptop o tablet")
    print("  2. Imprime las hojas que necesites o llévala en digital")
    print("  3. Recorre piso por piso. Marcaconciliado=NO si no aparece.")
    print("  4. Anota equipos NUEVOS en la hoja 'Pendientes_Nuevos'")
    print("  5. Avísale a Hermes al terminar para importar批量")
    print("=" * 60)


if __name__ == "__main__":
    main()
