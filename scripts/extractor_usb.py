#!/usr/bin/env python3
"""
extractor_usb.py v2 — Extracción automática desde carpeta 2026 del HGR No.1

CAMBIOS vs v1 (Opción A de OpenCode audit 2026-06-26):
- NO requiere columnas ruta_contrato / ruta_servicio en Excel
- Asocia PDFs por convención de carpetas:
  - MATENIMIENTOS DE EQUIPO MEDICO/CONTRATO CONSOLIDADO/*.pdf → tipo_adquisicion=contrato_consolidado
  - MATENIMIENTOS DE EQUIPO MEDICO/CONTRATO GARANTIA/*.pdf → tipo_adquisicion=garantia
  - MATENIMIENTOS DE EQUIPO MEDICO/RECURSO PROPIO/*.pdf → tipo_adquisicion=recurso_propio
  - Cualquier otra carpeta → tipo_adquisicion=subrogado
- Mapeo inteligente: nombre de archivo PDF contiene N/S o marca-modelo, se busca match
- Genera metadata.json con TODOS los equipos de TODOS los maestros

Uso:
    python3 extractor_usb.py /mnt/c/Users/djpiy/Desktop/Bioingeneria/2026 /tmp/contratos_migracion.zip

O sin args: detecta automáticamente la carpeta 2026 y crea el ZIP en /tmp/
"""
import os
import sys
import json
import zipfile
import re
import shutil
from pathlib import Path
from openpyxl import load_workbook

# Mapeo de carpeta -> tipo_adquisicion
TIPO_POR_CARPETA = {
    "CONTRATO CONSOLIDADO": "contrato_consolidado",
    "CONTRATO GARANTIA": "garantia",
    "RECURSO PROPIO": "recurso_propio",
    # todo lo demás → subrogado
}

# Patrones para extraer serie del nombre de archivo PDF
PATRON_SERIE = re.compile(r"[A-Z]{0,3}[\-\_]?\d{4,}[A-Z0-9]*", re.IGNORECASE)

# Headers conocidos que mapean a "serie" en distintos Excels
HEADERS_SERIE = ["serie", "no. serie", "n/s", "nº serie", "numero de serie", "no.serie", "n.serie", "serial"]

# Headers que mapean a "marca"
HEADERS_MARCA = ["marca"]

# Headers que mapean a "modelo"
HEADERS_MODELO = ["modelo"]

# Headers que mapean a "proveedor"
HEADERS_PROVEEDOR = ["proveedor", "empresa"]

# Headers que mapean a "numero_contrato_servicio"
HEADERS_CONTRATO = ["contrato", "no. de contrato", "numero de contrato", "no.contrato", "nº contrato", "no contrato"]

# Headers que mapean a "tipo_equipo"
HEADERS_TIPO = ["tipo equipo", "tipo_equipo", "equipo", "descripcion articulo completa", "descripción artículo completa"]


def detectar_header(headers, candidatos):
    """Busca cuál header del Excel matchea con uno de los candidatos. Retorna índice o -1."""
    for i, h in enumerate(headers):
        h_clean = str(h).strip().lower() if h else ""
        for c in candidatos:
            if c in h_clean:
                return i
    return -1


def leer_excel_maestro(path_excel):
    """Lee un Excel y devuelve lista de equipos extraídos (sin PDFs aún).

    Busca en todas las hojas del Excel, identifica columnas relevantes
    por nombre (case-insensitive) y extrae las filas con datos.

    MAX_ROWS_POR_HOJA = 5000 para evitar cuelgues con Excels de 1M+ filas
    (caso real: SEMAFORO DE MANTENIMIENTOS GRAVE.xlsx hoja CONSOLIDADO).
    """
    MAX_ROWS_POR_HOJA = 5000
    try:
        wb = load_workbook(path_excel, read_only=True, data_only=True)
    except Exception as e:
        print(f"  ERROR abriendo {path_excel}: {e}")
        return []

    equipos = []
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        if sheet.max_row < 2:
            continue
        # Detectar fila de headers (buscar la primera fila con varios strings no-vacíos)
        header_row = None
        for r in range(1, min(5, sheet.max_row + 1)):
            row_vals = [str(c.value).strip() if c.value else "" for c in sheet[r]]
            non_empty = sum(1 for v in row_vals if v and len(v) < 50)
            if non_empty >= 3:
                header_row = r
                break
        if header_row is None:
            continue
        headers = [str(c.value).strip().lower() if c.value else "" for c in sheet[header_row]]
        idx_serie = detectar_header(headers, HEADERS_SERIE)
        if idx_serie == -1:
            # Esta hoja no tiene serie, saltar
            continue
        idx_marca = detectar_header(headers, HEADERS_MARCA)
        idx_modelo = detectar_header(headers, HEADERS_MODELO)
        idx_proveedor = detectar_header(headers, HEADERS_PROVEEDOR)
        idx_contrato = detectar_header(headers, HEADERS_CONTRATO)
        idx_tipo = detectar_header(headers, HEADERS_TIPO)

        # Limitar a MAX_ROWS_POR_HOJA para evitar cuelgues con Excels de 1M+ filas
        max_row = min(header_row + 1 + MAX_ROWS_POR_HOJA, sheet.max_row + 1)
        if sheet.max_row + 1 > max_row:
            print(f"    ⚠️  {sheet_name}: {sheet.max_row} filas, limitando a {MAX_ROWS_POR_HOJA}")

        # Leer todas las filas del rango en una sola operación (read_only=True es eficiente)
        try:
            all_rows = list(sheet.iter_rows(min_row=header_row + 1, max_row=max_row - 1,
                                           min_col=1, max_col=sheet.max_column,
                                           values_only=True))
        except Exception as e:
            print(f"    ERROR leyendo {sheet_name}: {e}")
            continue

        for row in all_rows:
            if idx_serie >= len(row):
                continue
            serie = str(row[idx_serie]).strip() if row[idx_serie] else ""
            if not serie or serie in ("None", "N/A", "n/a", "-"):
                continue

            entry = {"serie": serie, "_source_file": os.path.basename(path_excel), "_source_sheet": sheet_name}
            if idx_marca >= 0 and idx_marca < len(row):
                v = row[idx_marca]
                if v: entry["marca"] = str(v).strip()
            if idx_modelo >= 0 and idx_modelo < len(row):
                v = row[idx_modelo]
                if v: entry["modelo"] = str(v).strip()
            if idx_proveedor >= 0 and idx_proveedor < len(row):
                v = row[idx_proveedor]
                if v: entry["proveedor_servicio"] = str(v).strip()
            if idx_contrato >= 0 and idx_contrato < len(row):
                v = row[idx_contrato]
                if v: entry["numero_contrato_servicio"] = str(v).strip()
            if idx_tipo >= 0 and idx_tipo < len(row):
                v = row[idx_tipo]
                if v: entry["nombre"] = str(v).strip()[:200]
            equipos.append(entry)
    wb.close()
    return equipos


def detectar_tipo_por_carpeta(ruta_pdf_rel):
    """Determina tipo_adquisicion según la carpeta donde está el PDF."""
    ruta_lower = ruta_pdf_rel.lower()
    for carpeta, tipo in TIPO_POR_CARPETA.items():
        if carpeta.lower() in ruta_lower:
            return tipo
    return "subrogado"


def buscar_pdfs_para_equipos(equipos, base_dir):
    """Para cada equipo, busca PDFs asociados en la carpeta 2026.

    Estrategia: buscar PDFs cuyo nombre contenga la serie del equipo
    O cuya carpeta indique el tipo_adquisicion.
    """
    # Indexar todos los PDFs con su tipo y serie inferida
    pdfs_por_tipo = {"contrato_consolidado": [], "garantia": [], "recurso_propio": [], "subrogado": []}
    for root, dirs, files in os.walk(base_dir):
        for f in files:
            if f.lower().endswith(".pdf"):
                full = os.path.join(root, f)
                rel = os.path.relpath(full, base_dir)
                tipo = detectar_tipo_por_carpeta(rel)
                pdfs_por_tipo[tipo].append((full, rel, f))

    print(f"  PDFs indexados por tipo: " + ", ".join(f"{k}={len(v)}" for k, v in pdfs_por_tipo.items()))

    # Asignar PDFs a equipos por matching de nombre
    # Para cada equipo, buscar PDF que contenga su serie en el nombre
    for eq in equipos:
        serie = eq["serie"]
        # Normalizar serie: "ABC-1234" → buscar "ABC1234" y "ABC-1234"
        serie_normalized = re.sub(r"[\s\-_/]", "", serie).lower()
        tipo = eq.get("_tipo_sugerido", "subrogado")

        # Buscar en PDFs del tipo sugerido primero
        candidatos = pdfs_por_tipo.get(tipo, []) + sum([pdfs_por_tipo[t] for t in pdfs_por_tipo if t != tipo], [])
        for full, rel, fname in candidatos:
            fname_normalized = re.sub(r"[\s\-_/]", "", fname).lower()
            if serie_normalized in fname_normalized or serie.lower() in fname.lower():
                eq["contrato_pdf_file"] = fname
                eq["_pdf_path"] = full
                eq["_pdf_rel"] = rel
                break

    return equipos


def agregar_tipo_sugerido(equipos, base_dir):
    """Para cada equipo, sugiere tipo_adquisicion basándose en la carpeta de su PDF."""
    for eq in equipos:
        rel = eq.get("_pdf_rel", "")
        if rel:
            eq["tipo_adquisicion"] = detectar_tipo_por_carpeta(rel)
        else:
            # Si no tiene PDF, intentar inferir por marca/modelo
            marca = eq.get("marca", "").lower()
            if "ge" in marca or "philips" in marca or "drager" in marca:
                eq["tipo_adquisicion"] = "contrato_consolidado"
            elif "hamilton" in marca or "mindray" in marca:
                eq["tipo_adquisicion"] = "garantia"
            else:
                eq["tipo_adquisicion"] = "recurso_propio"


def main():
    if len(sys.argv) >= 3:
        base_dir = sys.argv[1]
        output_zip = sys.argv[2]
    elif len(sys.argv) == 2:
        base_dir = sys.argv[1]
        output_zip = "/tmp/contratos_migracion.zip"
    else:
        # Auto-detect: buscar carpeta 2026 en ubicaciones comunes
        candidatos = [
            "/mnt/c/Users/djpiy/Desktop/Bioingeneria/2026",
            "/opt/sigab/USB_DATA/2026",
        ]
        base_dir = None
        for c in candidatos:
            if os.path.isdir(c):
                base_dir = c
                break
        if not base_dir:
            print("ERROR: no se encontró carpeta 2026. Especifica path como argumento.")
            sys.exit(1)
        output_zip = "/tmp/contratos_migracion.zip"

    print("=" * 60)
    print("EXTRACTOR USB v2 — HGR No.1 IMSS Tijuana")
    print("=" * 60)
    print(f"Base dir: {base_dir}")
    print(f"Output: {output_zip}")
    print()

    # Paso 1: identificar excels maestros
    print("[1/3] Leyendo Excels maestros...")
    masters_paths = [
        # Master #1: SEMAFORO (base, 421 series, 4 hojas)
        os.path.join(base_dir, "SEMAFORO DE MANTENIMIENTOS GRAVE.xlsx"),
        # Master #2: CEDULA CONSOLIDADO (149 series, 38 cols)
        os.path.join(base_dir, "MATENIMIENTOS DE EQUIPO MEDICO", "CONTRATO CONSOLIDADO", "CEDULA DE EQUIPO MEDICO CONTRATO CONSOLIDADO 2026 b.xlsx"),
        # Master #3: consolidados 2026 (146 series, con No. CONTRATO)
        os.path.join(base_dir, "MATENIMIENTOS DE EQUIPO MEDICO", "CONTRATO CONSOLIDADO", "consolidados 2026.xlsx"),
        # Master #4: CEDULA GARANTIA (115 series, con fechas)
        os.path.join(base_dir, "MATENIMIENTOS DE EQUIPO MEDICO", "CONTRATO GARANTIA", "CEDULA DE GARANTIAS 2026.xlsx"),
        # Auxiliares para series faltantes
        os.path.join(base_dir, "MATENIMIENTOS DE EQUIPO MEDICO", "CONTRATO GARANTIA", "CALENDARIO GARANTIAS INDIVIDUALES 2026.xlsx"),
        os.path.join(base_dir, "MATENIMIENTOS DE EQUIPO MEDICO", "RECURSO PROPIO", "CALENDARIO DE RECURSO PROPIO.xlsx"),
        os.path.join(base_dir, "MATENIMIENTOS DE EQUIPO MEDICO", "RECURSO PROPIO", "CALENDARIO INDIVIDUAL DE RECURSO PROPIO.xlsx"),
        os.path.join(base_dir, "EQUIPOS_GARANTIA_VENC_2026.xlsx"),
        os.path.join(base_dir, "Copia de LISTA_BAJA CALIFORNIA_2026.xlsx"),
        os.path.join(base_dir, "Copia de INVENTARIO VENTILADORES BCN.xlsx"),
        os.path.join(base_dir, "EQUIPOS QUE NO SE LES REALIZARA EL MANTENIMIENTO PREVENTIVO.xlsx"),
    ]
    # Filtrar los que existen
    masters_paths = [p for p in masters_paths if os.path.isfile(p)]
    print(f"  Masters encontrados: {len(masters_paths)}")
    for p in masters_paths:
        print(f"    - {os.path.relpath(p, base_dir)}")

    # Paso 2: leer todos los masters y consolidar
    print("\n[2/3] Extrayendo equipos de los masters...")
    equipos_total = []
    seen_series = set()
    for p in masters_paths:
        print(f"  Leyendo {os.path.basename(p)}...")
        equipos = leer_excel_maestro(p)
        # Deduplicar por serie (último gana)
        nuevos = 0
        duplicados = 0
        for eq in equipos:
            serie = eq["serie"]
            if serie in seen_series:
                duplicados += 1
            else:
                seen_series.add(serie)
                nuevos += 1
            equipos_total.append(eq)
        print(f"    → {nuevos} nuevos, {duplicados} duplicados (ya en otro master)")

    print(f"\n  Total equipos únicos: {len(seen_series)}")

    # Paso 3: buscar PDFs asociados
    print("\n[3/3] Buscando PDFs asociados...")
    equipos_total = buscar_pdfs_para_equipos(equipos_total, base_dir)
    agregar_tipo_sugerido(equipos_total, base_dir)

    # Limpiar campos internos antes de escribir
    equipos_limpios = []
    for eq in equipos_total:
        clean = {k: v for k, v in eq.items() if not k.startswith("_")}
        # Defaults razonables
        clean.setdefault("nombre", eq.get("serie", "?"))
        clean.setdefault("marca", "DESCONOCIDA")
        clean.setdefault("modelo", "DESCONOCIDO")
        clean.setdefault("ubicacion", "POR UBICAR")
        clean.setdefault("estado", "operativo")
        # Si tiene PDF y tipo_adquisicion, mantener
        if "contrato_pdf_file" in clean and "tipo_adquisicion" not in clean:
            clean["tipo_adquisicion"] = detectar_tipo_por_carpeta(eq.get("_pdf_rel", ""))
        equipos_limpios.append(clean)

    # Escribir ZIP
    metadata = {"contratos": equipos_limpios}
    print(f"\n  Generando ZIP: {output_zip}")
    with zipfile.ZipFile(output_zip, "w", zipfile.ZIP_DEFLATED) as zf:
        # metadata.json
        zf.writestr("metadata.json", json.dumps(metadata, indent=2, ensure_ascii=False))
        # PDFs
        pdfs_incluidos = set()
        for eq in equipos_limpios:
            pdf_path = eq.get("_pdf_path") or eq.get("_pdf_abs_path")
            if pdf_path and os.path.isfile(pdf_path) and pdf_path not in pdfs_incluidos:
                zf.write(pdf_path, os.path.basename(pdf_path))
                pdfs_incluidos.add(pdf_path)
    # Stats
    n_pdfs = sum(1 for eq in equipos_limpios if eq.get("contrato_pdf_file"))
    print(f"\n  Equipos en metadata.json: {len(equipos_limpios)}")
    print(f"  PDFs en ZIP: {len(pdfs_incluidos)}")
    print(f"  Tipos adquisición:")
    from collections import Counter
    tipos = Counter(eq.get("tipo_adquisicion", "?") for eq in equipos_limpios)
    for t, n in tipos.most_common():
        print(f"    {t}: {n}")
    print(f"\n  ZIP creado: {output_zip}")
    print(f"  Tamaño: {os.path.getsize(output_zip) // 1024} KB")
    print("\n" + "=" * 60)
    print("Siguientes pasos:")
    print("1. Subir el ZIP a SIGAB via UI o API")
    print("2. Verificar UPSERT con GET /api/equipos/?limit=10")
    print("3. Limpiar backups después de validar")
    print("=" * 60)


if __name__ == "__main__":
    main()
