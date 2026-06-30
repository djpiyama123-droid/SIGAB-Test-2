#!/usr/bin/env python3
"""
parse_2026.py — Parser de metadatos de la carpeta 2026 (HGR No.1) → staging.

Lee las CEDULAS/CALENDARIOS de mantenimiento y los nombres de .docx de
RECURSO PROPIO, normaliza series/inventarios y emite:
  - staging/equipos_2026.json   (un registro por equipo-fuente)
  - staging/equipos_2026.csv    (mismo contenido, plano)

NO toca la base de datos. El cruce contra `equipos` lo hace match_2026.py.

Mapeo carpeta -> tipo_adquisicion:
  CONTRATO CONSOLIDADO -> contrato_consolidado
  CONTRATO GARANTIA    -> garantia
  RECURSO PROPIO       -> recurso_propio
( 'subrogado' existe en el enum pero no hay carpeta fuente en 2026 )

Uso:
  python parse_2026.py [--root "/ruta/a/2026"] [--out staging]
"""
import argparse
import csv
import json
import os
import re
import sys
from pathlib import Path

import openpyxl

DEFAULT_ROOT = "/mnt/c/Users/djpiy/Desktop/Bioingeneria/2026"
MANT = "MATENIMIENTOS DE EQUIPO MEDICO"  # sic: typo en el nombre real de la carpeta

# ── normalización ──────────────────────────────────────────────────
_SEP = re.compile(r"[\s\-_./\\]+")


def norm(v):
    """Normaliza serie/inventario para match: upper, sin separadores."""
    if v is None:
        return ""
    s = str(v).strip().upper()
    s = _SEP.sub("", s)
    # descarta marcadores de vacío comunes en los Excel del IMSS
    if s in ("", "-", "NA", "N/A", "SN", "SINSERIE", "NI", "NINGUNO", "0"):
        return ""
    return s


def clean(v):
    """Texto legible: trim + colapsa espacios. None -> ''."""
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v)).strip()


def iso_date(v):
    """datetime/date -> 'YYYY-MM-DD'; texto -> tal cual limpio."""
    if v is None:
        return ""
    if hasattr(v, "strftime"):
        try:
            return v.strftime("%Y-%m-%d")
        except ValueError:
            return ""
    return clean(v)


# ── detección de cabecera ──────────────────────────────────────────
def find_header(ws, max_scan=15):
    """Devuelve (idx_fila_1based, {nombre_col_norm: idx}) de la fila que
    contenga una columna de serie. None si no se encuentra."""
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True), 1):
        cells = [clean(c).upper() for c in row]
        joined = " ".join(cells)
        if "SERIE" in joined and ("EQUIPO" in joined or "MARCA" in joined or "MODELO" in joined):
            cols = {}
            for idx, name in enumerate(cells):
                if name:
                    cols[name] = idx
            return i, cols
    return None


def col(cols, *needles):
    """Primer índice de columna cuyo nombre contenga alguno de los needles."""
    for name, idx in cols.items():
        for n in needles:
            if n in name:
                return idx
    return None


def cell(row, idx):
    if idx is None or idx >= len(row):
        return None
    return row[idx]


# ── parsers por fuente ─────────────────────────────────────────────
def parse_sheet(ws, tipo, archivo, hoja):
    """Extrae registros de una hoja con cabecera tipo cédula/calendario."""
    found = find_header(ws)
    if not found:
        return []
    hrow, cols = found
    c_serie = col(cols, "SERIE")
    c_inv = col(cols, "INVENTARIO")
    c_nom = col(cols, "EQUIPO", "NOMBRE")
    c_marca = col(cols, "MARCA")
    c_modelo = col(cols, "MODELO")
    c_prov = col(cols, "PROVEEDOR")
    c_precio = col(cols, "PRECIO")
    # nº de contrato REAL: evita "INICIO/FIN DE CONTRATO" (son año/fecha, no contrato)
    c_contrato = col(cols, "NO. CONTRATO", "NUMERO DE CONTRATO", "NÚMERO DE CONTRATO", "NUM CONTRATO", "No CONTRATO")
    # fechas / vigencia (heterogéneo entre cédulas)
    c_f1 = col(cols, "1ER MANTENIMIENTO", "PRIMER MANTENIMIEN")
    c_f2 = col(cols, "2DO MANTENIMIENTO", "SEGUNDO MANTENIMIE")
    c_ini = col(cols, "INICIO DE CONTRATO", "INICIO DE SEMANA 1")
    c_vig = col(cols, "ULTIMO MANTENIMIENTO", "VIGENCIA", "FIN DE SEMANA")

    out = []
    for row in ws.iter_rows(min_row=hrow + 1, values_only=True):
        serie_raw = clean(cell(row, c_serie))
        inv_raw = clean(cell(row, c_inv))
        nombre = clean(cell(row, c_nom))
        # cabecera repetida en medio de la hoja (frecuente en los calendarios IMSS)
        if serie_raw.upper() == "SERIE" or nombre.upper() in ("EQUIPO", "NOMBRE DEL EQUIPO"):
            continue
        # fila basura: sin serie ni inventario ni nombre util
        if not norm(serie_raw) and not norm(inv_raw):
            continue
        if not nombre and not norm(serie_raw):
            continue
        rec = {
            "tipo_adquisicion": tipo,
            "fuente": f"{archivo}::{hoja}",
            "serie_raw": serie_raw,
            "serie_norm": norm(serie_raw),
            "inventario_raw": inv_raw,
            "inventario_norm": norm(inv_raw),
            "nombre": nombre,
            "marca": clean(cell(row, c_marca)),
            "modelo": clean(cell(row, c_modelo)),
            "proveedor": clean(cell(row, c_prov)),
            "numero_contrato": clean(cell(row, c_contrato)),
            "precio_servicio": clean(cell(row, c_precio)),
            "fecha_1er_mant": iso_date(cell(row, c_f1)),
            "fecha_2do_mant": iso_date(cell(row, c_f2)),
            "inicio_contrato": iso_date(cell(row, c_ini)),
            "vigencia": iso_date(cell(row, c_vig)),
        }
        out.append(rec)
    return out


def parse_xlsx(path, tipo):
    recs = []
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        print(f"  ! ERROR abriendo {path.name}: {type(e).__name__}: {e}", file=sys.stderr)
        return recs
    for ws in wb.worksheets:
        recs += parse_sheet(ws, tipo, path.name, ws.title)
    wb.close()
    return recs


# ── parser de nombres .docx (RECURSO PROPIO) ───────────────────────
# patrones: "..._100032114920.docx", "... NS 0114098-01.docx", "..._2016LCM12008_2026.docx"
_INV_US = re.compile(r"_(\d[\d\-]{6,})(?:_\d{4})?$")          # _<digitos> opcional _2026
_SERIE_NS = re.compile(r"\bN[º°]?\.?\s*S\.?\s*([A-Z0-9][A-Z0-9\-]{3,})", re.I)
_RUIDO = re.compile(r"PORTADA|CAMBIO DE CELDAS|^LENTES$|OFTALMOSCOPIO|CLINICA DE MAMA|^AUTOCLAVE DE GAS", re.I)


def parse_docx_names(root, tipo="recurso_propio"):
    base = root / MANT / "RECURSO PROPIO"
    recs = []
    if not base.exists():
        return recs
    for p in base.rglob("*.docx"):
        name = p.name
        if name.startswith("~$"):
            continue
        stem = p.stem
        if _RUIDO.search(stem):
            # sin identificador fiable -> lo registramos como 'sin match' informativo
            recs.append({"tipo_adquisicion": tipo, "fuente": f"docx::{name}",
                         "serie_raw": "", "serie_norm": "", "inventario_raw": "", "inventario_norm": "",
                         "nombre": clean(stem), "marca": "", "modelo": "", "proveedor": "",
                         "numero_contrato": "", "precio_servicio": "", "fecha_1er_mant": "",
                         "fecha_2do_mant": "", "inicio_contrato": "", "vigencia": "", "_ruido": True})
            continue
        serie_raw = inv_raw = ""
        m = _SERIE_NS.search(stem)
        if m:
            serie_raw = m.group(1)
        m2 = _INV_US.search(stem)
        if m2:
            inv_raw = m2.group(1)
        recs.append({
            "tipo_adquisicion": tipo, "fuente": f"docx::{name}",
            "serie_raw": serie_raw, "serie_norm": norm(serie_raw),
            "inventario_raw": inv_raw, "inventario_norm": norm(inv_raw),
            "nombre": clean(stem), "marca": "", "modelo": "", "proveedor": "",
            "numero_contrato": "", "precio_servicio": "", "fecha_1er_mant": "",
            "fecha_2do_mant": "", "inicio_contrato": "", "vigencia": "",
        })
    return recs


# ── orquestación ───────────────────────────────────────────────────
SOURCES = [
    ("contrato_consolidado", "CONTRATO CONSOLIDADO", "CEDULA DE EQUIPO MEDICO CONTRATO CONSOLIDADO 2026 b.xlsx"),
    ("garantia", "CONTRATO GARANTIA", "CEDULA DE GARANTIAS 2026.xlsx"),
    ("recurso_propio", "RECURSO PROPIO", "CALENDARIO DE RECURSO PROPIO.xlsx"),
    ("recurso_propio", "RECURSO PROPIO", "CALENDARIO INDIVIDUAL DE RECURSO PROPIO.xlsx"),
]


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--root", default=DEFAULT_ROOT)
    ap.add_argument("--out", default=str(Path(__file__).parent / "staging"))
    args = ap.parse_args()
    root = Path(args.root)
    out = Path(args.out)
    out.mkdir(parents=True, exist_ok=True)

    if not root.exists():
        print(f"ERROR: no existe la carpeta raíz {root}", file=sys.stderr)
        sys.exit(2)

    recs = []
    for tipo, sub, fname in SOURCES:
        path = root / MANT / sub / fname
        if not path.exists():
            print(f"  ! falta fuente: {path}", file=sys.stderr)
            continue
        r = parse_xlsx(path, tipo)
        print(f"  + {fname[:45]:45} [{tipo:20}] -> {len(r)} filas")
        recs += r

    dn = parse_docx_names(root)
    print(f"  + nombres .docx RECURSO PROPIO            [recurso_propio]      -> {len(dn)} archivos")
    recs += dn

    # salida JSON
    (out / "equipos_2026.json").write_text(
        json.dumps(recs, ensure_ascii=False, indent=2), encoding="utf-8")
    # salida CSV
    cols = ["tipo_adquisicion", "serie_raw", "serie_norm", "inventario_raw", "inventario_norm",
            "nombre", "marca", "modelo", "proveedor", "numero_contrato", "precio_servicio",
            "fecha_1er_mant", "fecha_2do_mant", "inicio_contrato", "vigencia", "fuente"]
    with (out / "equipos_2026.csv").open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=cols, extrasaction="ignore")
        w.writeheader()
        for r in recs:
            w.writerow(r)

    # ── reporte ────────────────────────────────────────────────────
    by_tipo = {}
    con_serie = sin_serie = con_inv = ruido = 0
    series_norm = {}
    for r in recs:
        by_tipo[r["tipo_adquisicion"]] = by_tipo.get(r["tipo_adquisicion"], 0) + 1
        if r.get("_ruido"):
            ruido += 1
        if r["serie_norm"]:
            con_serie += 1
            series_norm[r["serie_norm"]] = series_norm.get(r["serie_norm"], 0) + 1
        else:
            sin_serie += 1
        if r["inventario_norm"]:
            con_inv += 1
    dups = {k: v for k, v in series_norm.items() if v > 1}

    print("\n" + "=" * 60)
    print(f"TOTAL registros:        {len(recs)}")
    for t, n in sorted(by_tipo.items()):
        print(f"  {t:22} {n}")
    print(f"con serie_norm:         {con_serie}")
    print(f"sin serie (revisar):    {sin_serie}  (de los cuales ruido docx: {ruido})")
    print(f"con inventario_norm:    {con_inv}")
    print(f"series_norm únicas:     {len(series_norm)}")
    print(f"series_norm duplicadas: {len(dups)}  {dict(list(dups.items())[:8])}")
    print(f"\nstaging -> {out/'equipos_2026.json'}\n          {out/'equipos_2026.csv'}")


if __name__ == "__main__":
    main()
